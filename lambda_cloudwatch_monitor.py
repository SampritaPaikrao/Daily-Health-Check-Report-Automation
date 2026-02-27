"""
AWS Lambda function to monitor EC2 and RDS CloudWatch metrics.
Collects CPU, Memory, and Disk utilization and flags resources exceeding 80% threshold.
Exports results to Excel file in S3.
"""

import json
import boto3
from datetime import datetime, timedelta
from io import BytesIO
import os
import base64
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

ses = boto3.client('ses', region_name='ap-south-2')

# Initialize AWS clients
cloudwatch = boto3.client('cloudwatch')
s3 = boto3.client('s3')

# Configuration
THRESHOLD = 80.0  # Percentage threshold for flagging
S3_BUCKET = os.environ.get('S3_BUCKET', 'lambda-packages-monitor-721')
METRIC_PERIOD = 300  # 5 minutes in seconds
LOOKBACK_MINUTES = 24 * 60  # 1440 minutes = 24 hours


def lambda_handler(event, context):
    """
    Main Lambda handler function.
    
    Expected event structure:
    {
        "servers": {
            "ec2_instances": [
                {"instance_id": "i-0229831aa95fcc6af", "name": "WebServer1"},
                {"instance_id": "i-0a81df9fad1de5d9b", "name": "AppServer1"}
            ],
            "rds_instances": [
                {"db_instance_id": "productiondb", "name": "productiondb"},
            ]
        }
    }
    """
    
    try:
        # Parse server details from event
        servers = event.get('servers', {})
        ec2_instances = servers.get('ec2_instances', [])
        rds_instances = servers.get('rds_instances', [])
        
        # Collect metrics
        results = []
        
        # Process EC2 instances
        for instance in ec2_instances:
            instance_id = instance.get('instance_id')
            name = instance.get('name', instance_id)
            
            print(f"Processing EC2 instance: {name} ({instance_id})")
            
            metrics = get_ec2_metrics(instance_id)
            results.append({
                'Type': 'EC2',
                'Name': name,
                'Resource_ID': instance_id,
                'CPU_Avg_24h': metrics['cpu']['avg'],
                'CPU_Max_24h': metrics['cpu']['max'],
                'CPU_Status': 'BAD' if metrics['cpu']['max'] > THRESHOLD else 'OK',
                'memory_Avg_24h': metrics['memory']['avg'],
                'memory_Max_24h': metrics['memory']['max'],
                'Memory_Status': 'BAD' if metrics['memory']['max'] > THRESHOLD else 'OK',
                'Disk_Avg_24h': metrics['disk']['avg'],
                'Disk_Max_24h': metrics['disk']['max'],
                'Disk_Status': 'BAD' if metrics['disk']['max'] > THRESHOLD else 'OK',
                'Overall_Status': 'BAD'
                    if (
                        metrics['cpu']['max'] > THRESHOLD or
                        metrics['memory']['max'] > THRESHOLD or
                        metrics['disk']['max'] > THRESHOLD
                    )
                    else 'OK',
                'Timestamp': datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
            })
        
        # Process RDS instances
        for instance in rds_instances:
            db_instance_id = instance.get('db_instance_id')
            name = instance.get('name', db_instance_id)
            
            print(f"Processing RDS instance: {name} ({db_instance_id})")
            
            metrics = get_rds_metrics(db_instance_id)
            results.append({
                'Type': 'RDS',
                'Name': name,
                'Resource_ID': db_instance_id,
                'CPU_Avg_24h': metrics['cpu']['avg'],
                'CPU_Max_24h': metrics['cpu']['max'],
                'CPU_Status': 'BAD' if metrics['cpu']['max'] > THRESHOLD else 'OK',
                'memory_Avg_24h': metrics['memory']['avg'],
                'memory_Max_24h': metrics['memory']['max'],
                'Memory_Status': 'BAD' if metrics['memory']['max'] > THRESHOLD else 'OK',
                'Disk_Avg_24h': metrics['disk']['avg'],
                'Disk_Max_24h': metrics['disk']['max'],
                'Disk_Status': 'BAD' if metrics['disk']['max'] > THRESHOLD else 'OK',
                'Overall_Status': 'BAD'
                    if (
                        metrics['cpu']['max'] > THRESHOLD or
                        metrics['memory']['max'] > THRESHOLD or
                        metrics['disk']['max'] > THRESHOLD
                    )
                    else 'OK',
                'Timestamp': datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
            })
        
        # Generate Excel file
        excel_buffer = create_excel_report(results)
        
        # Upload to S3
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        s3_key = f"cloudwatch-reports/metrics_report_{timestamp}.xlsx"
        
        s3.put_object(
            Bucket=S3_BUCKET,
            Key=s3_key,
            Body=excel_buffer.getvalue(),
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # SEND EMAIL 
        send_email_with_attachment(
            to_email="samprita9867@gmail.com",
            subject="Daily Health Check Report (EC2 & RDS)",
            body=f"""
            <h3>CloudWatch Monitoring Report</h3>
            <p>Last 24 hours EC2 & RDS utilization report attached.</p>
            <p><b>S3 Location:</b><br>
            s3://{S3_BUCKET}/{s3_key}</p>
            """,
            attachment_bytes=excel_buffer.getvalue(),
            filename=os.path.basename(s3_key)
)
        
        print(f"Report uploaded to s3://{S3_BUCKET}/{s3_key}")
        
        # Prepare response
        bad_resources = [r for r in results if r['Overall_Status'] == 'BAD']
        
        return {
            'statusCode': 200,
            'body': json.dumps({
                'message': 'Metrics collection completed successfully',
                'total_resources': len(results),
                'flagged_resources': len(bad_resources),
                's3_location': f"s3://{S3_BUCKET}/{s3_key}",
                'flagged_details': bad_resources
            })
        }
        
    except Exception as e:
        print(f"Error: {str(e)}")
        return {
            'statusCode': 500,
            'body': json.dumps({
                'message': 'Error processing metrics',
                'error': str(e)
            })
        }


def get_ec2_metrics(instance_id):
    """Get CloudWatch metrics for EC2 instance."""
    
    end_time = datetime.utcnow()
    start_time = end_time - timedelta(minutes=LOOKBACK_MINUTES)
    
    # CPU Utilization
    cpu = get_metric_stats(
        namespace='AWS/EC2',
        metric_name='CPUUtilization',
        dimensions=[{'Name': 'InstanceId', 'Value': instance_id}],
        start_time=start_time,
        end_time=end_time
    )
    
    # Memory Utilization (requires CloudWatch agent)
    memory = get_metric_stats(
        namespace='CWAgent',
        metric_name='mem_used_percent',
        dimensions=[{'Name': 'InstanceId', 'Value': instance_id}],
        start_time=start_time,
        end_time=end_time
    )
    
    # Disk Utilization (requires CloudWatch agent)
    disk = get_metric_stats(
        namespace='CWAgent',
        metric_name='disk_used_percent',
        dimensions=[
            {'Name': 'InstanceId', 'Value': instance_id},
            {'Name': 'path', 'Value': '/'},
        ],
        start_time=start_time,
        end_time=end_time
    )
    
    return {
        'cpu': cpu,
        'memory': memory,
        'disk': disk
    }


def get_rds_metrics(db_instance_id):
    """Get CloudWatch metrics for RDS instance."""
    
    end_time = datetime.utcnow()
    start_time = end_time - timedelta(minutes=LOOKBACK_MINUTES)
    
    # CPU Utilization
    cpu = get_metric_stats(
        namespace='AWS/RDS',
        metric_name='CPUUtilization',
        dimensions=[{'Name': 'DBInstanceIdentifier', 'Value': db_instance_id}],
        start_time=start_time,
        end_time=end_time
    )
    
    # Memory (Freeable Memory - convert to used percentage)
    freeable_memory = get_metric_stats(
        namespace='AWS/RDS',
        metric_name='FreeableMemory',
        dimensions=[{'Name': 'DBInstanceIdentifier', 'Value': db_instance_id}],
        start_time=start_time,
        end_time=end_time,
        statistic='Average'
    )
    
    # Get total memory from RDS instance details
    rds_client = boto3.client('rds')
    try:
        response = rds_client.describe_db_instances(DBInstanceIdentifier=db_instance_id)
        db_instance_class = response['DBInstances'][0]['DBInstanceClass']
        total_memory = get_rds_memory_by_instance_class(db_instance_class)
        
        if freeable_memory > 0 and total_memory > 0:
            memory = ((total_memory - freeable_memory) / total_memory) * 100
        else:
            memory = 0.0
    except:
        memory = 0.0
    
    # Disk Utilization (Free Storage Space - convert to used percentage)
    free_storage = get_metric_stats(
        namespace='AWS/RDS',
        metric_name='FreeStorageSpace',
        dimensions=[{'Name': 'DBInstanceIdentifier', 'Value': db_instance_id}],
        start_time=start_time,
        end_time=end_time
    )
    
    # Get allocated storage
    try:
        response = rds_client.describe_db_instances(DBInstanceIdentifier=db_instance_id)
        allocated_storage = response['DBInstances'][0]['AllocatedStorage'] * 1024 * 1024 * 1024  # Convert GB to bytes
        
        if free_storage > 0 and allocated_storage > 0:
            disk = ((allocated_storage - free_storage) / allocated_storage) * 100
        else:
            disk = 0.0
    except:
        disk = 0.0
    
    return {
        'cpu': cpu,
        'memory': memory,
        'disk': disk
    }


def get_metric_stats(namespace, metric_name, dimensions, start_time, end_time, statistic='Average'):
    """
    Returns avg, max and breach_count for last 24 hours
    """
    try:
        response = cloudwatch.get_metric_statistics(
            Namespace=namespace,
            MetricName=metric_name,
            Dimensions=dimensions,
            StartTime=start_time,
            EndTime=end_time,
            Period=METRIC_PERIOD,
            Statistics=['Average','Maximum']
        )

        datapoints = response.get('Datapoints', [])

        if not datapoints:
            print(f"No data for {namespace}/{metric_name}")
            return {"avg": 0.0, "max": 0.0, "breach_count": 0}

        avg_values = [dp['Average'] for dp in datapoints if 'Average' in dp]
        max_values = [dp['Maximum'] for dp in datapoints if 'Maximum' in dp]

        avg_val = sum(avg_values) / len(avg_values)
        max_val = max(max_values)
        breach_count = len([v for v in max_values if v > THRESHOLD])

        return {
            "avg": round(avg_val, 2),
            "max": round(max_val, 2),
            "breach_count": breach_count
        }

    except Exception as e:
        print(f"Error getting metric {namespace}/{metric_name}: {str(e)}")
        return {"avg": 0.0, "max": 0.0, "breach_count": 0}


def get_rds_memory_by_instance_class(instance_class):
    """
    Get approximate memory in bytes for RDS instance class.
    This is a simplified mapping - adjust based on your instance types.
    """
    
    memory_map = {
        'db.t3.micro': 1 * 1024 * 1024 * 1024,
        'db.t3.small': 2 * 1024 * 1024 * 1024,
        'db.t3.medium': 4 * 1024 * 1024 * 1024,
        'db.t3.large': 8 * 1024 * 1024 * 1024,
        'db.t3.xlarge': 16 * 1024 * 1024 * 1024,
        'db.t3.2xlarge': 32 * 1024 * 1024 * 1024,
        'db.m5.large': 8 * 1024 * 1024 * 1024,
        'db.m5.xlarge': 16 * 1024 * 1024 * 1024,
        'db.m5.2xlarge': 32 * 1024 * 1024 * 1024,
        'db.m5.4xlarge': 64 * 1024 * 1024 * 1024,
        'db.r5.large': 16 * 1024 * 1024 * 1024,
        'db.r5.xlarge': 32 * 1024 * 1024 * 1024,
        'db.r5.2xlarge': 64 * 1024 * 1024 * 1024,
    }
    
    return memory_map.get(instance_class, 8 * 1024 * 1024 * 1024)  # Default 8GB


def create_excel_report(results):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CloudWatch Metrics"

    #  FIXED HEADERS (ORDER + COMMAS)
    headers = [
        'Type',
        'Name',
        'Resource_ID',

        'CPU_Avg_24h (%)',
        'CPU_Max_24h (%)',
        'CPU_Status',

        'Memory_Avg_24h (%)',
        'Memory_Max_24h (%)',
        'Memory_Status',

        'Disk_Avg_24h (%)',
        'Disk_Max_24h (%)',
        'Disk_Status',

        'Overall_Status',
        'Timestamp'
    ]

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # WRITE ROWS (1-to-1 mapping)
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result['Type'])
        ws.cell(row=row, column=2, value=result['Name'])
        ws.cell(row=row, column=3, value=result['Resource_ID'])

        ws.cell(row=row, column=4, value=result.get('CPU_Avg_24h'))
        ws.cell(row=row, column=5, value=result.get('CPU_Max_24h'))
        ws.cell(row=row, column=6, value=result.get('CPU_Status'))

        ws.cell(row=row, column=7, value=result.get('memory_Avg_24h'))
        ws.cell(row=row, column=8, value=result.get('memory_Max_24h'))
        ws.cell(row=row, column=9, value=result.get('Memory_Status'))

        ws.cell(row=row, column=10, value=result.get('Disk_Avg_24h'))
        ws.cell(row=row, column=11, value=result.get('Disk_Max_24h'))
        ws.cell(row=row, column=12, value=result.get('Disk_Status'))

        ws.cell(row=row, column=13, value=result.get('Overall_Status'))
        ws.cell(row=row, column=14, value=result.get('Timestamp'))

        # COLOR Overall_Status
        status_cell = ws.cell(row=row, column=13)
        if status_cell.value == 'BAD':
            status_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            status_cell.font = Font(color='FFFFFF', bold=True)
        else:
            status_cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            status_cell.font = Font(bold=True)

    # Auto width
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def send_email_with_attachment(to_email, subject, body, attachment_bytes, filename):
    msg = MIMEMultipart()
    msg['Subject'] = subject
    msg['From'] = "samprita9867@gmail.com"
    msg['To'] = to_email

    msg.attach(MIMEText(body, 'html'))

    attachment = MIMEApplication(attachment_bytes)
    attachment.add_header(
        'Content-Disposition',
        'attachment',
        filename=filename
    )
    msg.attach(attachment)

    response = ses.send_raw_email(
        Source=msg['From'],
        Destinations=[to_email],
        RawMessage={'Data': msg.as_string()}
    )

    return response
